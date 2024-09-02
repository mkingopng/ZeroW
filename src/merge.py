import pandas as pd
import functools
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Define file paths relative to the script directory
file_paths = {
    "brisbane": os.path.join(script_dir, "brisbane.csv"),
    "southside": os.path.join(script_dir, "southside.csv"),
    "wales": os.path.join(script_dir, "wales.csv"),
    "gc": os.path.join(script_dir, "gold_coast.csv"),
    "cairns": os.path.join(script_dir, "cairns.csv"),
    "sunshine_coast": os.path.join(script_dir, "sunshine_coast.csv")
}

# Define locations corresponding to the file paths
locations = {
    "brisbane": "Brisbane",
    "southside": "South Side",
    "wales": "Wales",
    "gc": "Gold Coast",
    "cairns": "Cairns",
    "sunshine_coast": "Sunshine Coast"
}


def read_csv_with_error_handling(file_path, location):
    """
    Read a CSV file with error handling and add a 'Location' column.
    """
    if not os.path.exists(file_path):
        print(f"Error: File not found - {file_path}")
        return pd.DataFrame()  # Return an empty df if the file is not found
    try:
        df = pd.read_csv(file_path)
        df['Location'] = location
        return df
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return pd.DataFrame()

# Read the CSV files into DataFrames with location information
dfs = {name: read_csv_with_error_handling(path, locations[name]) for name, path in file_paths.items()}

# Rename 'CUSTOMER' to 'CUSTOMER_NAME' for Brisbane and Southside
for key in ['brisbane', 'southside']:
    if 'CUSTOMER' in dfs[key].columns:
        dfs[key].rename(columns={'CUSTOMER': 'CUSTOMER_NAME'}, inplace=True)

def standardize_column_names(df, new_columns):
    """
    Standardize column names of a DataFrame.
    """
    if not df.empty and len(df.columns) == len(new_columns):
        df.columns = new_columns
    else:
        print(f"Warning: Column length mismatch for {df.columns}. Expected {len(new_columns)} columns but got {len(df.columns)}. Skipping renaming.")
    return df

# Standardize column names
standardize_columns = {
    "wales": ["CUSTOMER_NAME", "EMAIL", "PHONE", "INTERVAL", "AGREED_DATE", "START_DATE", "STATUS", "Location"],
    "gc": ["CUSTOMER_NAME", "EMAIL", "PHONE", "INTERVAL", "AGREED_DATE", "START_DATE", "STATUS", "Location"],
    "cairns": ["AGREED_DATE", "START_DATE", "TYPE", "STATUS", "METHOD", "MERCHANT", "CUSTOMER_NAME", "REFERENCE", "PAYMENT_REQUEST", "NEXT_PAYMENT", "NEXT_PAYMENT_ON", "PRICE", "LAST_PAYMENT_ON", "CANCEL_DATE", "COMMENCEMENT_DATE", "COMPLETION_DATE", "TOTAL_VALUE", "PAID", "REMAINING", "AGREEMENT_ID", "Location"],
    "sunshine_coast": ["AGREED_DATE", "START_DATE", "TYPE", "STATUS", "METHOD", "MERCHANT", "CUSTOMER_NAME", "REFERENCE", "PAYMENT_REQUEST", "NEXT_PAYMENT", "NEXT_PAYMENT_ON", "PRICE", "LAST_PAYMENT_ON", "CANCEL_DATE", "COMMENCEMENT_DATE", "COMPLETION_DATE", "TOTAL_VALUE", "PAID", "REMAINING", "AGREEMENT_ID", "Location"],
    "brisbane": ["AGREED_DATE", "START_DATE", "TYPE", "STATUS", "METHOD", "CUSTOMER_NAME", "REFERENCE", "PAYMENT_REQUEST", "NEXT_PAYMENT", "NEXT_PAYMENT_ON", "PRICE", "LAST_PAYMENT_ON", "CANCEL_DATE", "COMMENCEMENT_DATE", "COMPLETION_DATE", "TOTAL_VALUE", "PAID", "REMAINING", "AGREEMENT_ID", "Location"],
    "southside": ["AGREED_DATE", "START_DATE", "TYPE", "STATUS", "METHOD", "CUSTOMER_NAME", "REFERENCE", "PAYMENT_REQUEST", "NEXT_PAYMENT", "NEXT_PAYMENT_ON", "PRICE", "LAST_PAYMENT_ON", "CANCEL_DATE", "COMMENCEMENT_DATE", "COMPLETION_DATE", "TOTAL_VALUE", "PAID", "REMAINING", "AGREEMENT_ID", "Location"]
}

for key, new_columns in standardize_columns.items():
    dfs[key] = standardize_column_names(dfs[key], new_columns)

# Ensure necessary columns like CUSTOMER_NAME exist in all DataFrames
for key in dfs.keys():
    if 'CUSTOMER_NAME' not in dfs[key].columns:
        print(f"Warning: CUSTOMER_NAME column missing in {key}.")
        dfs[key]['CUSTOMER_NAME'] = pd.NA

# Convert date columns to datetime objects
date_columns = [
    "AGREED_DATE", "START_DATE", "NEXT_PAYMENT_ON", "created",
    "current_period_start"
]


def convert_date_columns(df, date_columns):
    """
    Convert date columns to datetime objects.
    """
    for col in date_columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    return df


convert_date_columns_partial = functools.partial(convert_date_columns, date_columns=date_columns)
dfs = {name: convert_date_columns_partial(df) for name, df in dfs.items()}

# Concatenate DataFrames
merged_df = pd.concat(dfs.values(), ignore_index=True)  # concatenate dfs

# Drop unnecessary columns
columns_to_drop = [
    "EMAIL", "PHONE", "CITY", "STATE", "COUNTRY", "POSTCODE", "MOBILE",
    "AGREEMENT_ID"
]

merged_df.drop(columns=columns_to_drop, errors='ignore', inplace=True)

# Merge INTERVAL and REFERENCE into MEMBERSHIP_TYPE
merged_df['MEMBERSHIP_TYPE'] = merged_df['INTERVAL'].combine_first(merged_df['REFERENCE'])

# Ensure MEMBERSHIP_TYPE exists and initialize it with NaN
if 'MEMBERSHIP_TYPE' not in merged_df.columns:
    merged_df['MEMBERSHIP_TYPE'] = pd.NA

# Remove any duplicate columns
merged_df = merged_df.loc[:, ~merged_df.columns.duplicated()]

# Mapping of PRICE to MEMBERSHIP_TYPE
price_to_membership = {
    35: "Standard",
    50: "Silver",
    99: "Gold",
    179: "Platinum",
    70: "Standard Plus Nutrition",
    85: "Silver Plus Nutrition",
    134: "Gold Plus Nutrition",
    214: "Platinum Plus Nutrition"
}


def impute_membership_type(row, price_to_membership):
    """
    Impute missing MEMBERSHIP_TYPE values based on PRICE.
    """
    if pd.isna(row['MEMBERSHIP_TYPE']):
        price = row['PRICE'] if 'PRICE' in row else None
        return price_to_membership.get(price, "Other")
    return row['MEMBERSHIP_TYPE']


# Impute missing MEMBERSHIP_TYPE
impute_membership_type_partial = functools.partial(impute_membership_type, price_to_membership=price_to_membership)
merged_df['MEMBERSHIP_TYPE'] = merged_df.apply(impute_membership_type_partial, axis=1)

# Normalize MEMBERSHIP_TYPE values
membership_type_mapping = {
    "Gold": ["Gold", "GOLD", "Gold Membership", "ZeroW Gold"],
    "Platinum": ["PLATINUM", "Platinum", "Platinum Membership"],
    "Standard": ["STANDARD membership", "Standard ZeroW Membe", "Standard", "Standard Memberhip", "ZeroW Standard"],
    # Add other mappings as needed
}


def normalize_membership_type(membership_type, mapping):
    """
    Clean up MEMBERSHIP_TYPE values.
    """
    membership_type = str(membership_type).strip()
    for standard, variants in mapping.items():
        if membership_type in variants:
            return standard
    return membership_type


normalize_membership_type_partial = functools.partial(normalize_membership_type, mapping=membership_type_mapping)
merged_df['MEMBERSHIP_TYPE'] = merged_df['MEMBERSHIP_TYPE'].apply(normalize_membership_type_partial)

# Normalize STATUS values
merged_df['STATUS'] = merged_df['STATUS'].str.strip().str.capitalize()

merged_df = merged_df[~merged_df['STATUS'].isin(['Cancelled', 'Canceled'])]

merged_df = merged_df[merged_df['NEXT_PAYMENT'] != 0]

# Drop unnecessary columns
merged_df.drop(columns=[
    'INTERVAL', 'MERCHANT', 'REFERENCE', 'PAYMENT_REQUEST', 'METHOD', 'PAYMENT_REQUEST', 'NEXT_PAYMENT',
    'NEXT_PAYMENT_ON', 'PRICE', 'COMMENCEMENT_DATE', 'COMPLETION_DATE', 'TOTAL_VALUE',
    'PAID', 'REMAINING'
], inplace=True)


with pd.ExcelWriter('merged_data.xlsx') as writer:
    merged_df.to_excel(writer, sheet_name="merged_data", index=False)

# Load the workbook and select the worksheet
wb = load_workbook('merged_data.xlsx')

# Iterate through all sheets in the workbook
for sheet in wb.sheetnames:
    ws = wb[sheet]
    ws.auto_filter.ref = ws.dimensions  # add an auto filter to all columns

wb.save('merged_data.xlsx')  # Save the workbook with filters applied

wb.close()

print('complete')

# todo: upgrade to database to store data rather than excel
